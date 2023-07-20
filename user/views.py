from django.shortcuts import render
from Core.serializers import AssetsSerializer
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from rest_framework.permissions import IsAuthenticated, AllowAny
from Core.models import Assets,User
from .serializers import *
from rest_framework.pagination import LimitOffsetPagination, PageNumberPagination
from django.utils import timezone
from django.db import transaction
from Core.serializers import CompanyLoginSerializer
from rest_framework_simplejwt.tokens import RefreshToken
from rest_framework_simplejwt.authentication import JWTAuthentication
from django.db.models import Sum
from openpyxl import Workbook
from django.http import HttpResponse
import calendar
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from django.shortcuts import get_object_or_404
from rest_framework.permissions import BasePermission
from rest_framework_simplejwt.authentication import JWTAuthentication
from romulus_admin.custom_permissions import IsAdmin,IsUser

from romulus_admin.common_views import *



# Create your views here.
class AssetsView(AssetsCommonView):
    permission_classes = [IsUser]
    
 

class StaffAPIView(APIView):
    permission_classes = [AllowAny]

    def post(self, request):
        serializer = StaffSerializer(data=request.data)
        print(request.data)
        if serializer.is_valid():
            serializer.save()
            return Response({'message': 'Staff added successfully'})
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


    def get(self, request, id):
        staff_only = request.query_params.get('staff_only')
        if staff_only :
            staff_members = User.objects.filter(company_id=id,role='staff',is_active=True).order_by('username')
        else:
            staff_members = User.objects.filter(company_id=id,is_active=True).order_by('username')
        serializer = GetStaffSerializer(staff_members, many=True)
        return Response(serializer.data)
    
    def put(self, request, id):
        user = get_object_or_404(User, pk=id)
        serializer = StaffSerializer(user, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save()
            return Response({'message': 'Staff updated successfully'})
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
    def patch(self, request, id):
        instance = get_object_or_404(User, id=id)
        password = request.data.get('password')
        if password:
            instance.password = make_password(password)
            instance.save()
            return Response({'message': 'Password changed successfully'})
        return Response({'message': 'No password provided'}, status=status.HTTP_400_BAD_REQUEST)
    
    def delete(self, request, id):
        user = User.objects.get(id=id)
        user.delete()
        return Response({'Deleted Sucessfully'},status=status.HTTP_200_OK)
    

# class OrderAPIView(APIView):
#     permission_classes = [AllowAny]

#     def post(self,request):
#         serializer = OrderSerializer(data=request.data)
#         if serializer.is_valid():
#             serializer.save()
#             return Response({'message': 'Order placed successfully'})


class OrderView(OrderCommonView):
    permission_classes = [IsUser]

class GetDieselPrice(GetDieselPrice):
    permission_classes = [IsUser]

class MyPaginator(PageNumberPagination):
    page_size = 12
    page_size_query_param = 'page_size'
    max_page_size = 100


class OrderHistoryAPIView(OrderHistoryCommonView):
    permission_classes = [IsUser]




class TransactionsView(TransactionsCommonView):
    permission_classes = [IsUser]

    


class DashboardView(APIView):
    permission_classes = [IsUser]
    def get(self,request,id):
        try : 
            current_month = datetime.now().month
            current_year = datetime.now().year

            print('qqqqq',id)
            total_price = Order.objects.filter(company=id,order_status='Delivered',created_at__year=current_year, created_at__month=current_month).aggregate(total_price=Sum('total_price'))['total_price']
            total_quantity = Order.objects.filter(company=id,order_status='Delivered',created_at__year=current_year, created_at__month=current_month).aggregate(total_quantity=Sum('quantity'))['total_quantity']

            monthly_saved_amt = Order.objects.filter(company=id,order_status='Delivered',created_at__year=current_year, created_at__month=current_month).aggregate(saved_amount=Sum('saved_amount'))['saved_amount']
            total_saved_amt = Order.objects.filter(company=id,order_status='Delivered').aggregate(saved_amount=Sum('saved_amount'))['saved_amount']
            company_obj = CompanyInfo.objects.get(company=id)

            dashboard_data = {
                'monthly_purchase_cost' : total_price,
                'monthly_purchase_quantity' : total_quantity,
                'total_outstanding': company_obj.total_outstanding,
                'monthly_saved_amt':monthly_saved_amt,
                'total_saved_amt':total_saved_amt
            }

            return Response(data=dashboard_data,status=status.HTTP_200_OK)
        
        except:
            return Response('Something went wrong',status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class ExportOrdersView(APIView):
    permission_classes = [AllowAny] 

    def get(self, request):
        # Query the orders data from the database
        company_id = request.query_params.get('company_id')
        filter_type = request.query_params.get('filter_type')
        
        
        if filter_type == 'date_range':
            from_date = request.query_params.get('from_date')
            to_date = request.query_params.get('to_date')
            from_date = timezone.make_aware(datetime.strptime(from_date, '%Y-%m-%d'))
            to_date = timezone.make_aware(datetime.strptime(to_date, '%Y-%m-%d'))
            orders = Order.objects.filter(company=company_id, created_at__range=[from_date, to_date])
        else:
            month = request.query_params.get('month')
            month_number = list(calendar.month_name).index(month.capitalize())
            orders = Order.objects.filter(company=company_id, created_at__month=month_number)

        # Create a new Excel workbook and worksheet
        wb = Workbook()
        ws = wb.active

        # Apply formatting to the headers
        header_font = Font(bold=True)
        for col_num, header in enumerate(['Order ID', 'Ordered By','User Type', 'Asset','Quantity', 'Total Price',  'Created At'], start=1):
            col_letter = get_column_letter(col_num)
            cell = ws['{}1'.format(col_letter)]
            cell.value = header
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')

        # Apply formatting to the order data
        for row_num, order in enumerate(orders, start=2):
            ws.cell(row=row_num, column=1, value=order.id)
            ws.cell(row=row_num, column=2, value=order.ordered_by.username)
            ws.cell(row=row_num, column=3, value=order.ordered_user_type)
            ws.cell(row=row_num, column=4, value=order.asset.assetName)
            ws.cell(row=row_num, column=5, value=order.quantity)
            ws.cell(row=row_num, column=6, value=order.total_price)
            ws.cell(row=row_num, column=7, value=order.created_at.astimezone(timezone.utc).replace(tzinfo=None))

        # Auto-size columns
        for col_num in range(1, 8):
            col_letter = get_column_letter(col_num)
            ws.column_dimensions[col_letter].auto_size = True
        
        ws.column_dimensions['G'].width = 20

        # Create the HTTP response with the Excel file content
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=orders.xlsx'

        # Save the workbook to the response
        wb.save(response)

        return response





# class UserLoginView(APIView):
#     permission_classes = [AllowAny]
#     def post(self, request):
#         serializer = CompanyLoginSerializer(data=request.data)
#         if serializer.is_valid():
#             company = serializer.validated_data['company']
#             refresh_token = RefreshToken.for_user(company)

#             response_data = {
#                 'message': 'Login successful',
#                 'user': {
#                     'id': company.id,
#                     'username': company.username,
#                     # Include any other user information you need
#                 }
#             }
#             res= Response()
#             res.set_cookie(
#             key='access_token',
#             value=str(refresh_token.access_token),
#             httponly=True,  
#             # samesite='None',   
#             secure=True
#             )

#             res.set_cookie(
#             key='refresh_token',
#             value=str(refresh_token),
#             httponly=True,
#             samesite='None',
#             secure=True
#             )
            
#             res.data = str(refresh_token.access_token)
#             res["X-CSRFToken"] = csrf.get_token(request)
#             return res

#         return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)



class CheckAuthView(APIView):
 
    permission_classes = [IsUser]
   
    def get(self, request):
        user = request.user
        print('fff',request.user)
        if user.role != 'company':
            try:
                company_name = User.objects.get(id=user.company_id_id).username
            except:
                company_name = 'Unknown'
        else:
            company_name = user.username
        response_data = {
            'message': 'Authentication successful',
            'user': {
                'id': user.id,
                'username': user.username,
                'user_type': user.role,
                'company_id':user.company_id_id,
                'company_name':company_name
                # Include any other user information you need
            }
        }
        return Response(response_data,status=status.HTTP_200_OK)


class TotalizerView(APIView):
    def post(self, request):
        print('ggggggg')
        serializer = TotalizerReadingsSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(200)
        print(serializer.errors)
        return Response(500)





class PopulateOrder(APIView):
    def get(self,request):
        company = User.objects.get(id=34)
        asset = Assets.objects.get(id=20)
        for i in range(50):
            order = Order.objects.create(
                company=company,
                ordered_by=company,
                quantity=10,
                asset=asset,
                diesel_price=3.5,
                total_price=35.0,
                order_status='ordered',
                created_at=timezone.now(),
                ordered_user_type='manager'
            )
        return Response(200)